import logging
from datetime import datetime, date
from os.path import exists
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pandas import DataFrame
import pandas as pd

from config import PathsConfig, AppInsightsConfig
from models.FeedbackModel import FeedbackModel
from models.NotificationTemplate import NAME, USER_ID, WORKSPACE_ID, MESSAGE_1
from models.Parameters import Parameters
from services.appInsightsService import AppInsightsService
from services.fileService import get_file_name, get_month_starter


class ExcelService:
    log = logging.getLogger(__name__)

    def append_data_to_file(self, file: str, data: DataFrame, param: Parameters):
        self.log.debug(f'appending {len(data)} records to {file}')

        try:
            writer = self.get_writer(file)
            data.to_excel(writer, startrow=writer.sheets[param.WorkSheet].max_row, index=False, header=False)
            writer.save()

        except BaseException as ex:
            self.log.error(f'exception during appending data, {ex}')

    def create_excel_file(self, data: DataFrame, param: Parameters, file_name: str, template_name: str):
        try:
            file = PathsConfig.LOCAL_FILE_PATH + file_name
            template_file = load_workbook(PathsConfig.TEMPLATE_FILE_PATH + template_name)

            writer = pd.ExcelWriter(file, engine='openpyxl')
            writer.book = template_file
            writer.save()

            self.append_data_to_file(file=file, data=data, param=param)

        except BaseException as ex:
            self.log.error(f'exception during creating file, {ex}')

    def prepare_excel(self, data: DataFrame, paths: PathsConfig, params: Parameters):

        prev_month_updated = False
        cur_month_updated = False
        cur_month_created = False

        today = date.today()
        start_date = pd.to_datetime(get_month_starter(), utc=True)

        prev_month = data.loc[(pd.to_datetime(data['timestamp'], utc=True) < start_date)]
        cur_month = data.loc[(pd.to_datetime(data['timestamp'], utc=True) >= start_date)]

        prev_month.insert(0, 'status', 'new')
        cur_month.insert(0, 'status', 'new')
        sub_path = paths.LOCAL_FILE_PATH + params.ReportSubPath

        if len(prev_month.values) > 0:
            file = sub_path + get_file_name(month=today.month - 1)
            self.append_data_to_file(file, prev_month, param=params)
            self.add_data_validations(file, params)
            prev_month_updated = True

        if len(cur_month.values) > 0:
            file = sub_path + get_file_name()
            if exists(file):
                self.append_data_to_file(file, cur_month, param=params)
                cur_month_updated = True
            else:
                file_name = params.ReportSubPath + get_file_name()
                self.create_excel_file(cur_month, params, file_name, params.ReportTemplate)
                cur_month_created = True
                self.log.info(f'created new file {get_file_name()}')

            self.add_data_validations(file, params)

        return prev_month_updated, cur_month_updated, cur_month_created

    def get_application_insights_data(self, config: AppInsightsConfig, start_date: datetime, query: str):
        client = AppInsightsService(config.AZURE_APPLICATION_INSIGHTS_BASE_URL, config.AZURE_APPLICATION_INSIGHTS_ID,
                                    config.AZURE_APPLICATION_INSIGHTS_KEY)

        self.log.debug('getting data from application insights')
        df = client.query_as_df(query, start_date=start_date)

        self.log.debug('getting data finished')
        return df

    def prepare_notifications_file(self, file_list: list, params: Parameters):
        try:
            feedback_df = pd.DataFrame()
            feedback_model_list = []
            for file in file_list:
                # model = FeedbackModel(file.Id, file.Name)
                model = {
                    'FieldId': file.Id,
                    "FileName": file.Name,
                    "QuestionIds": []
                }
                data = pd.read_excel(PathsConfig.LOCAL_FILE_PATH + params.ReportSubPath + file.Name, sheet_name=params.WorkSheet)
                feedback = data.loc[data['status'] == params.FeedbackSendMark]

                for index, item in feedback.iterrows():
                    feedback_data = DataFrame({
                        NAME: [item.user],
                        USER_ID: [item.userId],
                        WORKSPACE_ID: [item.workspaceId],
                        MESSAGE_1: ['predefined message goes here']
                    })
                    model['QuestionIds'].append(item.id)
                    feedback_df = pd.concat([feedback_df, feedback_data], axis=0)

                feedback_model_list.append(model)
            self.create_excel_file(feedback_df, params, params.NotificationList, params.NotificationTemplate)
            return feedback_model_list

        except BaseException as ex:
            self.log.error(f'exception during prepare_notifications_file, {ex}')

    def complete_send_feedback(self, feedback_delivered: [], params: Parameters):
        folder_path = PathsConfig.LOCAL_FILE_PATH + params.ReportSubPath
        try:
            for item in feedback_delivered:
                file = folder_path + item['FileName']
                wb = load_workbook(file)
                ws = wb[params.WorkSheet]
                for index in range(1, ws.max_row + 1):
                    if ws[f'B{index}'].value in item['QuestionIds']:
                        ws[f'A{index}'].value = params.FeedbackCompleteMark
                wb.save(file)
                self.add_data_validations(file, params)

        except BaseException as ex:
            self.log.error(f'exception during complete_send_feedback, {ex}')
        print()

    def add_data_validations(self, file: str, param: Parameters):
        try:
            writer = self.get_writer(file)

            dv = DataValidation(type="list", formula1=param.Formula, allow_blank=False, showErrorMessage=True)
            dv.add(param.FormulaRange)
            writer.sheets[param.WorkSheet].add_data_validation(dv)

            writer.save()
        except BaseException as ex:
            self.log.error(f'exception during adding data validation data, {ex}')

    def get_writer(self, file: str):
        try:
            book = load_workbook(file)
            writer = pd.ExcelWriter(file, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            return writer

        except BaseException as ex:
            self.log.error(f'cant create excel writer {ex}')
