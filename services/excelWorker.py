import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from models.FeedbackModel import FeedbackModel

template_path = './data/templates/'
data_path = './data/'


def crete_excel_file_from_template():
    book = load_workbook('services/template.xlsx')
    data = pd.DataFrame({
        'id': ['id1', 'id2', 'Amol', 'Lini'],
        'status': ['new', 'in_review', 'asdasd', 1],
        'user': ['Somu', 'user1', 'Amol', 'user54'],
        'timestamp': ['today', 'Kiku', 'Amol', 'Lini'],
        'userId': ['00001', '1231231', '1111', 'Lini']
    })
    writer = pd.ExcelWriter('new file.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    try:
        dv = DataValidation(type="list", formula1='=fields!$A$2:$A$9', allow_blank=False, showErrorMessage=True)

        dv.add('A2:A1048576')

        '=fields!$A$2:$A$9'
        worksheet = writer.sheets['Sheet1']
        worksheet.add_data_validation(dv)
    except BaseException as err:

        print("An exception occurred")

    data.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)

    writer.save()


def process_files(file_list: list):
    try:

        feedback_df = pd.DataFrame()
        feedback_model_list = []
        for file in file_list:
            data = pd.read_excel(data_path + 'reports/' + file['file_name'], sheet_name='Sheet1')
            feedback = data.loc[data['status'] == 'send_feedback']
            for index, item in feedback.iterrows():
                user = item['user']
                feedback_data = DataFrame({
                    'user name': [user],
                    'user email': [f'{user} @nike.com'],
                    'message 1': ['costom message 1'],
                    'message2': ['custom message 2']
                })
                feedback_df = pd.concat([feedback_df, feedback_data], axis=0)
                model = FeedbackModel(file['file_id'], file['file_name'], item['id'])
                feedback_model_list.append(model)

        feedback_df.to_excel()

        return feedback_model_list
    except BaseException as exc:
        print(exc)


def send_feedback(data: DataFrame):
    print(data)


def prepare_feedback_template(data: DataFrame, file: any):
    template_file = template_path + 'feedback_template.xlsx'
    feedback_df = pd.DataFrame()
    book = load_workbook(template_file)
    writer = pd.ExcelWriter(data_path + 'notifications/' + file['file_name'], engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for index, item in data.iterrows():
        user = item['user']
        feedback_data = DataFrame({
            'user name': [user],
            'user email': [f'{user} @nike.com'],
            'message 1': ['costom message 1'],
            'message2': ['custom message 2']
        })
        feedback_df = pd.concat([feedback_df, feedback_data], axis=0)
    feedback_df.to_excel(writer, startrow=1, index=False, header=False)
    writer.save()

    print()
